from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from modules.schedule.text_parser import ParsedActivityText, parse_activity_text


@dataclass
class ActivityBlock:
    """
    单个活动在 Excel 中对应的合并块信息。

    输入：
    - row: 活动块所在行号（1 基）
    - start_col: 起始列号（1 基）
    - end_col: 结束列号（1 基）
    - raw_value: 单元格原始文本内容

    输出：
    - 作为后续解析负责人、活动名与时间段的基础元信息。
    """

    row: int
    start_col: int
    end_col: int
    raw_value: Any


def _get_structure_cfg(config: Dict[str, Any]) -> Dict[str, Any]:
    return config.get("structure") or {}


def _get_datetime_cfg(config: Dict[str, Any]) -> Dict[str, Any]:
    return config.get("datetime") or {}


def _get_text_rules(config: Dict[str, Any]) -> Dict[str, Any]:
    return config.get("text_rules") or {}


def _get_duration_note_cfg(config: Dict[str, Any]) -> Dict[str, Any]:
    return config.get("duration_note") or {}


def _build_col_date_mapping(ws: Worksheet, config: Dict[str, Any]) -> Dict[int, date]:
    """
    根据第 1 行（日期行）构建“列号 → 日期”映射。

    输入：
    - ws: openpyxl 工作表对象
    - config: 配置字典，使用 structure.header_row_index / structure.data_start_col_index
             与 datetime.date_format / datetime.default_year。

    输出：
    - 字典：列号（int）到日期对象（date）的映射。

    逻辑：
    - 从 data_start_col_index 开始，读取 header_row_index 行上的值；
    - 若为 datetime/date 类型，直接取 date 部分；
    - 若为字符串，则按 date_format 解析，缺失年份时使用 default_year。
    """
    structure_cfg = _get_structure_cfg(config)
    dt_cfg = _get_datetime_cfg(config)

    header_row = int(structure_cfg.get("header_row_index", 1))
    data_start_col = int(structure_cfg.get("data_start_col_index", 3))
    date_format = dt_cfg.get("date_format", "%m/%d")
    default_year = int(dt_cfg.get("default_year", datetime.today().year))

    col_to_date: Dict[int, date] = {}
    for cell in ws[header_row]:
        if cell.column < data_start_col:
            continue
        value = cell.value
        if value is None:
            continue
        if isinstance(value, datetime):
            col_to_date[cell.column] = value.date()
        elif isinstance(value, date):
            col_to_date[cell.column] = value
        else:
            # 统一转为字符串按配置解析
            text = str(value).strip()
            if not text:
                continue
            try:
                dt = datetime.strptime(text, date_format)
                col_to_date[cell.column] = date(
                    year=default_year,
                    month=dt.month,
                    day=dt.day,
                )
            except Exception:
                # 若解析失败则跳过该列
                continue
    return col_to_date


def _build_row_mappings(ws: Worksheet, config: Dict[str, Any]) -> Tuple[Dict[int, str], Dict[int, str]]:
    """
    构建“行号 → 项目 / 活动类别”映射。

    输入：
    - ws: 工作表对象
    - config: 配置字典，使用 structure.project_col_index / structure.category_col_index。

    输出：
    - (row_to_project, row_to_category) 两个字典。

    逻辑：
    - 通过 merged_cells.ranges 识别第 1 列与第 2 列的纵向合并块；
    - 将合并块覆盖的所有行映射到同一个项目 / 活动类别；
    - 对于未合并但单独一行的情况，也根据单元格值补充映射。
    """
    structure_cfg = _get_structure_cfg(config)
    project_col = int(structure_cfg.get("project_col_index", 1))
    category_col = int(structure_cfg.get("category_col_index", 2))

    row_to_project: Dict[int, str] = {}
    row_to_category: Dict[int, str] = {}

    # 先处理纵向合并块
    for merged_range in ws.merged_cells.ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col

        if min_col == project_col:
            value = ws.cell(row=min_row, column=project_col).value
            if value is None:
                continue
            text = str(value).strip()
            if not text:
                continue
            for r in range(min_row, max_row + 1):
                row_to_project[r] = text
        elif min_col == category_col:
            value = ws.cell(row=min_row, column=category_col).value
            if value is None:
                continue
            text = str(value).strip()
            if not text:
                continue
            for r in range(min_row, max_row + 1):
                row_to_category[r] = text

    # 再补充未合并的单元格
    max_row = ws.max_row
    for r in range(1, max_row + 1):
        if r not in row_to_project:
            value = ws.cell(row=r, column=project_col).value
            if value is not None and str(value).strip():
                row_to_project[r] = str(value).strip()
        if r not in row_to_category:
            value = ws.cell(row=r, column=category_col).value
            if value is not None and str(value).strip():
                row_to_category[r] = str(value).strip()

    return row_to_project, row_to_category


# 用于过滤 activity_name 为纯“x天”的正则（与 config 无关，避免 YAML 转义或字符差异）
_ACTIVITY_NAME_ONLY_DURATION_RE = re.compile(r"^\d+\s*天$")


def _is_activity_name_only_duration(activity_name: Any) -> bool:
    """
    判断活动名称是否仅为“x天”形式（如 7天、9天），此类行将不写入结果。

    输入：
    - activity_name: 解析得到的活动名称，可为 str 或 None。
    输出：
    - True 表示应删除该行（不写入记录）；False 表示保留。
    逻辑：
    - 仅当整段名称匹配“数字+可选空格+天”时返回 True；
    - "love hearts（7天）" 等不匹配，返回 False。
    """
    if activity_name is None:
        return False
    s = str(activity_name).strip()
    return bool(_ACTIVITY_NAME_ONLY_DURATION_RE.fullmatch(s))


def _is_duration_note(text: str, config: Dict[str, Any]) -> bool:
    """
    判断给定文本是否为“n天”类备注。

    输入：
    - text: 原始文本
    - config: 配置字典，使用 duration_note.pattern 与 duration_note.enabled。

    输出：
    - bool: True 表示应视为“n天”备注，需要在解析活动时跳过。
    """
    duration_cfg = _get_duration_note_cfg(config)
    if not duration_cfg.get("enabled", True):
        return False
    pattern = duration_cfg.get("pattern") or ""
    if not pattern:
        return False
    try:
        regex = re.compile(pattern)
    except Exception:
        return False
    return bool(regex.fullmatch(text.strip()))


def _iter_merged_activity_blocks(
    ws: Worksheet,
    config: Dict[str, Any],
    used_cells: Dict[Tuple[int, int], bool],
) -> Iterable[ActivityBlock]:
    """
    遍历工作表中的合并单元格，生成活动块候选。

    输入：
    - ws: 工作表对象
    - config: 配置字典
    - used_cells: 记录属于合并区域的单元格坐标，用于后续识别非合并单元格活动。

    输出：
    - 迭代返回 ActivityBlock 实例，已过滤掉明显的“n天”备注块。
    """
    structure_cfg = _get_structure_cfg(config)
    header_row = int(structure_cfg.get("header_row_index", 1))
    empty_row = int(structure_cfg.get("empty_row_index", 2))
    data_start_row = max(header_row, empty_row) + 1
    data_start_col = int(structure_cfg.get("data_start_col_index", 3))

    for merged_range in ws.merged_cells.ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col

        # 记录该合并块覆盖的所有单元格坐标
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                used_cells[(r, c)] = True

        # 只关心数据区域内、且在活动区的横向合并块
        if min_row < data_start_row or min_col < data_start_col:
            continue

        value = ws.cell(row=min_row, column=min_col).value
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue

        # 过滤“n天”这类备注块
        if _is_duration_note(text, config):
            continue

        yield ActivityBlock(
            row=min_row,
            start_col=min_col,
            end_col=max_col,
            raw_value=value,
        )


def _iter_single_cell_activity_blocks(
    ws: Worksheet,
    config: Dict[str, Any],
    used_cells: Dict[Tuple[int, int], bool],
) -> Iterable[ActivityBlock]:
    """
    遍历非合并单元格区域，识别单日活动块。

    输入：
    - ws: 工作表对象
    - config: 配置字典
    - used_cells: 所有已归属合并块的单元格坐标

    输出：
    - 迭代返回 ActivityBlock，单元格的 start_col 与 end_col 相同。
    """
    structure_cfg = _get_structure_cfg(config)
    header_row = int(structure_cfg.get("header_row_index", 1))
    empty_row = int(structure_cfg.get("empty_row_index", 2))
    data_start_row = max(header_row, empty_row) + 1
    data_start_col = int(structure_cfg.get("data_start_col_index", 3))

    for row in ws.iter_rows(
        min_row=data_start_row,
        max_row=ws.max_row,
        min_col=data_start_col,
        max_col=ws.max_column,
    ):
        for cell in row:
            coord = (cell.row, cell.column)
            if used_cells.get(coord):
                continue
            value = cell.value
            if value is None:
                continue
            text = str(value).strip()
            if not text:
                continue
            if _is_duration_note(text, config):
                continue
            yield ActivityBlock(
                row=cell.row,
                start_col=cell.column,
                end_col=cell.column,
                raw_value=value,
            )


def _iter_activity_blocks(ws: Worksheet, config: Dict[str, Any]) -> Iterable[ActivityBlock]:
    """
    统一遍历工作表中的活动块（合并与非合并单元格）。

    输入：
    - ws: 工作表对象
    - config: 配置字典

    输出：
    - 迭代返回 ActivityBlock 实例。
    """
    used_cells: Dict[Tuple[int, int], bool] = {}
    # 先处理合并块
    for block in _iter_merged_activity_blocks(ws, config, used_cells):
        yield block
    # 再处理单个非合并单元格
    for block in _iter_single_cell_activity_blocks(ws, config, used_cells):
        yield block


def _parse_single_sheet(
    ws: Worksheet,
    sheet_name: str,
    config: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """
    解析单个工作表，返回该 sheet 内的所有活动记录。

    输入：
    - ws: 工作表对象
    - sheet_name: 工作表名称，用于可选的元数据输出
    - config: 解析配置字典

    输出：
    - 字典列表，每个字典是一条活动记录，字段包括：
      project, category, activity_name, owner, start_date, end_date, sheet_name(可选)、time_note(可选)
    """
    col_to_date = _build_col_date_mapping(ws, config)
    if not col_to_date:
        return []

    row_to_project, row_to_category = _build_row_mappings(ws, config)
    text_rules = _get_text_rules(config)

    structure_cfg = _get_structure_cfg(config)
    data_start_col = int(structure_cfg.get("data_start_col_index", 3))

    output_meta = config.get("output_meta") or {}
    add_sheet_name = bool(output_meta.get("add_sheet_name", False))
    sheet_name_field = output_meta.get("sheet_name_field", "sheet_name")

    output_cfg = config.get("output") or {}
    time_note_field = output_cfg.get("time_note_field")

    records: List[Dict[str, Any]] = []
    for block in _iter_activity_blocks(ws, config):
        # 需要能从行号映射到项目和类别
        project = row_to_project.get(block.row, "").strip()
        category = row_to_category.get(block.row, "").strip()
        if not project and not category:
            # 若既无项目又无类别信息，则跳过该活动块
            continue

        # 根据列号映射起止日期
        start_col = max(block.start_col, data_start_col)
        end_col = max(block.end_col, start_col)
        start_date = col_to_date.get(start_col)
        end_date = col_to_date.get(end_col)
        if start_date is None or end_date is None:
            # 若无法从列头推断日期，则跳过
            continue

        parsed: ParsedActivityText = parse_activity_text(
            raw_value=block.raw_value,
            default_start_date=start_date,
            default_end_date=end_date,
            datetime_cfg=_get_datetime_cfg(config),
            text_rules=text_rules,
        )

        # 若活动名称整列为“x天”形式（如 "7天"、"9天"），视为时长备注而非真实活动，直接跳过。
        # 不删除如 "love hearts（7天）" 这类名称（仅当 activity_name 整体匹配 数字+天 时才删）。
        if _is_activity_name_only_duration(parsed.activity_name):
            continue

        rec: Dict[str, Any] = {
            "project": project,
            "category": category,
            "activity_name": parsed.activity_name,
            "owner": parsed.owner,
            "start_date": parsed.start_date,
            "end_date": parsed.end_date,
        }
        # 计算活动天数：end_date - start_date + 1
        if parsed.start_date is not None and parsed.end_date is not None:
            rec["activity_days"] = (parsed.end_date - parsed.start_date).days + 1

        if time_note_field and parsed.time_note:
            rec[time_note_field] = parsed.time_note
        if add_sheet_name:
            rec[sheet_name_field] = sheet_name
        records.append(rec)

    return records


def parse_schedule_file(file_path: Path, config: Dict[str, Any]) -> pd.DataFrame:
    """
    解析包含多个 Sheet 的活动排期 Excel 文件，输出合并后的活动明细表。

    输入：
    - file_path: 排期表文件路径（Path 对象）
    - config: 从 configs/schedule_parser.yaml 读取的解析配置字典。

    输出：
    - pandas.DataFrame：多 sheet 记录合并后的总表。

    逻辑：
    - 根据 sheets.names 与 sheets.name_pattern 选择需要解析的工作表；
    - 对每个工作表调用 _parse_single_sheet 生成活动记录；
    - 将所有记录合并为 DataFrame，若不存在记录则返回空 DataFrame。
    """
    if not file_path.exists():
        raise FileNotFoundError(f"排期表文件不存在：{file_path}")

    wb = load_workbook(filename=file_path, data_only=True)

    sheets_cfg = config.get("sheets") or {}
    name_list: List[str] = sheets_cfg.get("names") or []
    name_pattern: str = sheets_cfg.get("name_pattern") or ""

    # 决定需要解析的 sheet 名称
    target_sheet_names: List[str] = []
    if name_list:
        target_sheet_names = [name for name in name_list if name in wb.sheetnames]
    elif name_pattern:
        import re

        try:
            pattern = re.compile(name_pattern)
            target_sheet_names = [name for name in wb.sheetnames if pattern.search(name)]
        except re.error:
            target_sheet_names = list(wb.sheetnames)
    else:
        # 若未配置，则默认解析所有 sheet
        target_sheet_names = list(wb.sheetnames)

    all_records: List[Dict[str, Any]] = []
    for sheet_name in target_sheet_names:
        ws = wb[sheet_name]
        sheet_records = _parse_single_sheet(ws, sheet_name, config)
        all_records.extend(sheet_records)

    if not all_records:
        return pd.DataFrame(
            columns=[
                "project",
                "category",
                "activity_name",
                "owner",
                "start_date",
                "end_date",
                "activity_days",
            ]
        )

    return pd.DataFrame(all_records)

